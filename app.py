import os
import re
import tempfile
from io import BytesIO

import pandas as pd
import pytesseract
import streamlit as st
from pdf2image import convert_from_bytes
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="PDF Table Extractor", page_icon="📄", layout="wide")


# -----------------------------
# OCR + parsing helpers
# -----------------------------
def ocr_page(image: Image.Image, lang: str = "eng", psm: int = 6) -> str:
    config = f"--oem 3 --psm {psm}"
    return pytesseract.image_to_string(image, lang=lang, config=config)


def ocr_tsv(image: Image.Image, lang: str = "eng", psm: int = 6) -> pd.DataFrame:
    config = f"--oem 3 --psm {psm}"
    df = pytesseract.image_to_data(
        image,
        lang=lang,
        config=config,
        output_type=pytesseract.Output.DATAFRAME
    )
    df = df.dropna(subset=["text"])
    df["text"] = df["text"].astype(str).str.strip()
    df = df[df["text"] != ""]
    return df


def group_words_into_lines(tsv_df: pd.DataFrame, y_threshold: int = 12):
    """
    Group OCR words into text lines using approximate y-position clustering.
    Returns list of lines, each line is list of word dicts sorted by x.
    """
    if tsv_df.empty:
        return []

    words = []
    for _, row in tsv_df.iterrows():
        words.append({
            "text": str(row["text"]).strip(),
            "left": int(row["left"]),
            "top": int(row["top"]),
            "width": int(row["width"]),
            "height": int(row["height"]),
            "conf": float(row["conf"]) if str(row["conf"]).replace(".", "", 1).isdigit() else -1
        })

    words.sort(key=lambda w: (w["top"], w["left"]))

    lines = []
    current_line = []
    current_y = None

    for word in words:
        if current_y is None:
            current_line = [word]
            current_y = word["top"]
        elif abs(word["top"] - current_y) <= y_threshold:
            current_line.append(word)
        else:
            current_line.sort(key=lambda w: w["left"])
            lines.append(current_line)
            current_line = [word]
            current_y = word["top"]

    if current_line:
        current_line.sort(key=lambda w: w["left"])
        lines.append(current_line)

    return lines


def merge_words(line_words):
    return " ".join(w["text"] for w in sorted(line_words, key=lambda x: x["left"])).strip()


def split_line_into_columns(line_words, page_width: int, num_columns: int):
    """
    Heuristic column assignment using word x-position.
    Works best when table columns are vertically aligned.
    """
    col_width = page_width / num_columns
    cols = [[] for _ in range(num_columns)]

    for word in line_words:
        idx = min(int(word["left"] / col_width), num_columns - 1)
        cols[idx].append(word["text"])

    return [" ".join(c).strip() for c in cols]


def looks_like_header(row_values, headers):
    row_text = " ".join(row_values).lower()
    header_hits = 0
    for h in headers:
        parts = [p for p in re.split(r"[\s/()'-]+", h.lower()) if p]
        if any(p in row_text for p in parts[:2]):
            header_hits += 1
    return header_hits >= max(2, len(headers) // 3)


def clean_rows(rows, headers):
    cleaned = []
    for row in rows:
        if not any(cell.strip() for cell in row):
            continue
        if looks_like_header(row, headers):
            continue
        cleaned.append(row[:len(headers)] + [""] * max(0, len(headers) - len(row)))
    return cleaned


def extract_tables_from_pdf(
    pdf_bytes: bytes,
    headers,
    dpi=250,
    lang="eng",
    psm=6,
    skip_first_pages=0,
    skip_last_pages=0,
    combine_multiline_rows=False,
    serial_column_index=0
):
    images = convert_from_bytes(pdf_bytes, dpi=dpi)
    total_pages = len(images)

    start_idx = skip_first_pages
    end_idx = total_pages - skip_last_pages if skip_last_pages > 0 else total_pages
    selected_images = images[start_idx:end_idx]

    all_rows = []
    progress_logs = []

    sno_pattern = re.compile(r"^\d{1,4}[.)-]?$")

    previous_row = None

    for page_no, img in enumerate(selected_images, start=start_idx + 1):
        progress_logs.append(f"OCR page {page_no}/{total_pages}")
        tsv_df = ocr_tsv(img, lang=lang, psm=psm)
        lines = group_words_into_lines(tsv_df)

        page_rows = []
        page_width = img.width

        for line_words in lines:
            row = split_line_into_columns(line_words, page_width, len(headers))

            if not any(c.strip() for c in row):
                continue

            if combine_multiline_rows and previous_row is not None:
                first_cell = row[serial_column_index].strip()
                if not sno_pattern.match(first_cell):
                    for i in range(len(headers)):
                        if row[i].strip():
                            if previous_row[i].strip():
                                previous_row[i] += " " + row[i].strip()
                            else:
                                previous_row[i] = row[i].strip()
                    continue

            page_rows.append(row)
            previous_row = row

        all_rows.extend(page_rows)

    all_rows = clean_rows(all_rows, headers)
    df = pd.DataFrame(all_rows, columns=headers)
    return df, progress_logs


# -----------------------------
# Excel builder
# -----------------------------
def build_excel_workbook(df: pd.DataFrame, sheet_name: str, title: str) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    title_fill = PatternFill("solid", fgColor="2E75B6")
    alt_fill = PatternFill("solid", fgColor="DCE6F1")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="FFFFFF", size=13)

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    ncols = len(df.columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    c.border = border
    ws.row_dimensions[1].height = 24

    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 28

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = border
            cell.alignment = wrap
        ws.row_dimensions[row_idx].height = 32

    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) if v is not None else 0 for v in df.iloc[:, i - 1].tolist()]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------
# UI
# -----------------------------
st.title("📄 PDF Table Extractor")
st.caption("Convert scanned PDF tables to Excel using OCR (Tesseract)")

with st.sidebar:
    st.header("Settings")

    headers_input = st.text_area(
        "Column headers (one per line)",
        value="S.No\nName of Candidate\nFather's/Husband's Name\nPermanent Address\nBlock Name\nUDISE Code\nPOP\nDistrict",
        height=220
    )

    ocr_lang = st.text_input("OCR language", value="eng")
    dpi = st.slider("PDF to image DPI", 150, 400, 250, 10)
    psm = st.selectbox("Tesseract PSM", [4, 6, 11, 12], index=1)
    skip_first_pages = st.number_input("Skip first pages", min_value=0, value=0, step=1)
    skip_last_pages = st.number_input("Skip last pages", min_value=0, value=0, step=1)
    combine_multiline_rows = st.checkbox("Merge multiline rows", value=True)

    sheet_name = st.text_input("Excel sheet name", value="Extracted Table")
    report_title = st.text_input("Excel title", value="Extracted PDF Table")

uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

if uploaded_file:
    headers = [h.strip() for h in headers_input.splitlines() if h.strip()]

    if len(headers) < 2:
        st.error("Please provide at least 2 column headers.")
    else:
        col1, col2 = st.columns([1, 1])

        with col1:
            if st.button("Extract Table", type="primary"):
                try:
                    with st.spinner("Running OCR and extracting table..."):
                        pdf_bytes = uploaded_file.read()

                        df, logs = extract_tables_from_pdf(
                            pdf_bytes=pdf_bytes,
                            headers=headers,
                            dpi=dpi,
                            lang=ocr_lang,
                            psm=psm,
                            skip_first_pages=skip_first_pages,
                            skip_last_pages=skip_last_pages,
                            combine_multiline_rows=combine_multiline_rows,
                            serial_column_index=0
                        )

                        st.session_state["extracted_df"] = df
                        st.session_state["logs"] = logs

                    st.success(f"Done. Extracted {len(df)} rows.")

                except Exception as e:
                    st.exception(e)

        with col2:
            if st.button("Clear Output"):
                st.session_state.pop("extracted_df", None)
                st.session_state.pop("logs", None)
                st.rerun()

if "logs" in st.session_state:
    with st.expander("Processing log", expanded=False):
        for line in st.session_state["logs"]:
            st.write(line)

if "extracted_df" in st.session_state:
    df = st.session_state["extracted_df"]

    st.subheader("Preview")
    st.dataframe(df, use_container_width=True, height=500)

    excel_data = build_excel_workbook(
        df=df,
        sheet_name=sheet_name,
        title=report_title
    )

    st.download_button(
        label="⬇️ Download Excel",
        data=excel_data,
        file_name="extracted_table.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    csv_data = df.to_csv(index=False).encode("utf-8")
    st.download_button(
        label="⬇️ Download CSV",
        data=csv_data,
        file_name="extracted_table.csv",
        mime="text/csv"
    )

st.markdown("---")
st.markdown(
    """
**Tips**
- Best for scanned PDFs with visible table columns.
- If rows are broken across multiple lines, keep **Merge multiline rows** enabled.
- Increase **DPI** for poor-quality scans.
- Try **PSM 4** or **PSM 11** if extraction looks messy.
- For Hindi or mixed-language OCR, install matching Tesseract language packs and use `eng+hin`.
"""
)
